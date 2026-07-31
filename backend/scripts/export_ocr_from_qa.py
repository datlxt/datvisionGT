"""Export a CLEAN fast-plate-ocr training set from QA workbooks + their videos.

For each ``<name>_qa.xlsx`` it pairs ``<name>_video.mp4``, reads the human "Biển số đúng"
column with the "Từ"/"Đến" window, then re-extracts plate crops in that window. Each crop
is kept ONLY if the base OCR read is within an edit distance of the human label — this
removes the label noise that ruined the first fine-tune (wide windows caught OTHER
vehicles/cards under the wrong label) while KEEPING the valuable hard cases (D->J, H->W)
that are only 1-2 characters off.

Vietnamese EV plates use "Đ" (e.g. 89MĐ701678) which the model alphabet lacks, so Đ is
mapped to D (the plate reads as ...MD...). Output matches fast-plate-ocr: an ``images/``
folder plus ``train.csv`` / ``val.csv`` with columns ``image_path,plate_text``.

    docker compose exec worker python -m scripts.export_ocr_from_qa \
        --input-dir /app/storage/qa_trainset \
        --output-root /app/storage/datasets/vn-ocr-finetune \
        --frames-per-plate 25 --max-edit-distance 3
"""

from __future__ import annotations

import argparse
import csv
import random
import re
from pathlib import Path

import cv2
import openpyxl

from app.core.config import get_settings
from app.vision.plate.domain import _plate_edit_distance
from app.vision.plate.fastalpr_adapter import FastAlprPlateEngine, crop_bgr


def _norm(text: object) -> str:
    """Uppercase alnum only; map Vietnamese Đ/đ -> D (model alphabet has no Đ)."""

    # upper() maps đ(U+0111) -> Đ(U+0110); then map Đ -> D (alphabet has no Đ).
    s = str(text).upper().replace("Đ", "D")
    return re.sub(r"[^A-Z0-9]", "", s)


def _valid_plate(plate: str) -> bool:
    # Accept normal (29D128130) and EV-as-D (89MD701678) VN plates.
    return bool(re.fullmatch(r"\d{2}[A-Z][A-Z0-9]\d{3,6}", plate))


def _parse_time_ms(value: object) -> int | None:
    if value is None:
        return None
    parts = str(value).strip().split(":")
    if not parts or parts[-1] == "":
        return None
    try:
        seconds = float(parts[-1])
        minutes = float(parts[-2]) if len(parts) >= 2 else 0.0
        hours = float(parts[-3]) if len(parts) >= 3 else 0.0
    except ValueError:
        return None
    return int((hours * 3600 + minutes * 60 + seconds) * 1000)


def _column_index(header: tuple, *names: str) -> int | None:
    for name in names:
        for index, cell in enumerate(header):
            text = str(cell).strip().lower() if cell is not None else ""
            if name.lower() in text:
                return index
    return None


def _plate_like(bbox: tuple[int, int, int, int]) -> bool:
    width, height = bbox[2] - bbox[0], bbox[3] - bbox[1]
    return width >= 40 and height >= 20 and 0.9 <= width / height <= 2.6


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--frames-per-plate", type=int, default=25)
    parser.add_argument("--sample-every-ms", type=int, default=150)
    parser.add_argument("--max-edit-distance", type=int, default=3)
    parser.add_argument("--val-ratio", type=float, default=0.15)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    settings = get_settings()
    root = settings.model_root.resolve()
    engine = FastAlprPlateEngine(
        root / settings.plate_detector_model_path,
        root / settings.plate_ocr_model_path,
        root / settings.plate_ocr_config_path,
    )
    images_dir = args.output_root / "images"
    images_dir.mkdir(parents=True, exist_ok=True)

    # Group samples by plate text so the val split holds out WHOLE plates (no leakage).
    per_plate: dict[str, list[tuple[str, str]]] = {}
    kept = rejected = 0
    for qa_path in sorted(args.input_dir.glob("*_qa.xlsx")):
        base = qa_path.name[: -len("_qa.xlsx")]
        video = args.input_dir / f"{base}_video.mp4"
        if not video.is_file():
            print(f"skip (no video): {qa_path.name}")
            continue
        workbook = openpyxl.load_workbook(qa_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        header = rows[0]
        col_from = _column_index(header, "từ", "start", "from")
        col_to = _column_index(header, "đến", "end", "to")
        col_gt = _column_index(header, "biển số đúng", "gt", "expected")
        if None in (col_from, col_to, col_gt):
            print(f"skip (columns not found): {qa_path.name}")
            continue

        capture = cv2.VideoCapture(str(video))
        for row in rows[1:]:
            if col_gt >= len(row):
                continue
            plate = _norm(row[col_gt] or "")
            if not _valid_plate(plate):
                continue
            start_ms = _parse_time_ms(row[col_from]) if col_from < len(row) else None
            end_ms = _parse_time_ms(row[col_to]) if col_to < len(row) else None
            if start_ms is None or end_ms is None or end_ms < start_ms:
                continue
            saved = 0
            for timestamp in range(start_ms, end_ms + 1, args.sample_every_ms):
                if saved >= args.frames_per_plate:
                    break
                capture.set(cv2.CAP_PROP_POS_MSEC, timestamp)
                ok, frame = capture.read()
                if not ok:
                    continue
                detections = engine.detect(frame)
                if not detections:
                    continue
                best = max(detections, key=lambda d: d.confidence)
                if not _plate_like(best.bbox):
                    continue
                try:
                    crop = crop_bgr(frame, best.bbox)
                except ValueError:
                    continue
                # Keep only if the crop actually IS this plate (base OCR within edit distance).
                reading = engine._read(crop)
                if reading is None:
                    rejected += 1
                    continue
                if _plate_edit_distance(_norm(reading.raw_text), plate) > args.max_edit_distance:
                    rejected += 1
                    continue
                name = f"{base}_{plate}_{timestamp}.jpg"
                cv2.imwrite(str(images_dir / name), crop)
                per_plate.setdefault(plate, []).append((f"images/{name}", plate))
                saved += 1
                kept += 1
        capture.release()
        print(f"{qa_path.name}: plates so far {len(per_plate)}, crops kept {kept}")

    # Hold out whole plates for validation to avoid the same plate in train+val.
    plates = list(per_plate)
    random.Random(args.seed).shuffle(plates)
    n_val = max(1, int(len(plates) * args.val_ratio))
    val_plates = set(plates[:n_val])
    train = [s for p in plates if p not in val_plates for s in per_plate[p]]
    val = [s for p in val_plates for s in per_plate[p]]
    for csv_name, out_rows in (("train.csv", train), ("val.csv", val)):
        with (args.output_root / csv_name).open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["image_path", "plate_text"])
            writer.writerows(out_rows)
    print(
        f"\nDONE: kept {kept} crops (rejected {rejected} noisy) from {len(plates)} plates "
        f"-> train={len(train)} val={len(val)}"
    )


if __name__ == "__main__":
    main()
