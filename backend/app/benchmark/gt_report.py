"""Event-level GT evaluator for the motorcycle plate report.

Compares model events against a human QA ground-truth workbook (the ``Lane9_qa.xlsx``
layout: ``From - To`` + ``License Plate expected`` + quality + ``Result``) and reports
detection precision/recall, OCR exact-match / CER, and a per-quality breakdown.

The metric core (:func:`evaluate_events`) is pure and dependency-free so it can be unit
tested and reused from a CLI, a worker, or an API without pulling in Excel or the DB.
Workbook parsing is isolated in the ``parse_*`` helpers and only needs ``openpyxl``.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from app.vision.plate.domain import normalize_vietnamese_plate

NO_PLATE = "LPN_NO_PLATE_VEHICLE"


@dataclass(frozen=True, slots=True)
class GtEvent:
    index: int
    start_ms: int
    end_ms: int
    expected_plate: str  # normalized plate, or the NO_PLATE sentinel
    quality: str = ""
    human_result: str = ""  # TP/FP as labelled by the QA reviewer, when present


@dataclass(frozen=True, slots=True)
class ModelEvent:
    start_ms: int
    end_ms: int
    plate: str  # normalized plate, NO_PLATE sentinel, or "" when unreadable


@dataclass(frozen=True, slots=True)
class MatchedPair:
    gt: GtEvent
    model: ModelEvent
    overlap_ms: int
    plate_correct: bool


@dataclass(frozen=True, slots=True)
class EvalReport:
    detection: dict[str, float]
    recognition: dict[str, float]
    by_quality: dict[str, dict[str, int]]
    matched: list[MatchedPair] = field(default_factory=list)
    missed: list[GtEvent] = field(default_factory=list)
    extra: list[ModelEvent] = field(default_factory=list)


def normalize_plate(text: str | None) -> str:
    value = (text or "").strip()
    if value.upper() == NO_PLATE:
        return NO_PLATE
    return normalize_vietnamese_plate(value)


def _overlap_ms(a: GtEvent, b: ModelEvent) -> int:
    return max(0, min(a.end_ms, b.end_ms) - max(a.start_ms, b.start_ms))


def _edit_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    previous = list(range(len(right) + 1))
    for i, left_char in enumerate(left, start=1):
        current = [i]
        for j, right_char in enumerate(right, start=1):
            current.append(
                min(
                    current[-1] + 1,
                    previous[j] + 1,
                    previous[j - 1] + (left_char != right_char),
                )
            )
        previous = current
    return previous[-1]


def _ratio(numerator: float, denominator: float) -> float:
    return float(numerator / denominator) if denominator else 0.0


def evaluate_events(
    gt_events: list[GtEvent],
    model_events: list[ModelEvent],
    *,
    min_overlap_ms: int = 1,
) -> EvalReport:
    """Greedy time-overlap matching, then detection + recognition metrics.

    Each model event matches at most one GT event (highest overlap first). Unmatched GT
    events are misses (FN); unmatched model events are extras (FP) — this is exactly how
    a duplicate track fragment surfaces as a precision loss.
    """

    pairs = sorted(
        (
            (_overlap_ms(gt, model), gt_i, model_i)
            for gt_i, gt in enumerate(gt_events)
            for model_i, model in enumerate(model_events)
            if _overlap_ms(gt, model) >= min_overlap_ms
        ),
        reverse=True,
    )
    gt_taken: dict[int, int] = {}
    model_taken: set[int] = set()
    for _, gt_i, model_i in pairs:
        if gt_i in gt_taken or model_i in model_taken:
            continue
        gt_taken[gt_i] = model_i
        model_taken.add(model_i)

    matched: list[MatchedPair] = []
    by_quality: dict[str, dict[str, int]] = {}
    char_errors = char_total = readable_matched = exact = 0

    for gt_i, gt in enumerate(gt_events):
        quality = gt.quality or "unknown"
        bucket = by_quality.setdefault(quality, {"total": 0, "matched": 0, "correct": 0})
        bucket["total"] += 1
        model_i = gt_taken.get(gt_i)
        if model_i is None:
            continue
        model = model_events[model_i]
        correct = model.plate == gt.expected_plate and gt.expected_plate != ""
        matched.append(MatchedPair(gt=gt, model=model, overlap_ms=_overlap_ms(gt, model),
                                   plate_correct=correct))
        bucket["matched"] += 1
        if correct:
            bucket["correct"] += 1
            exact += 1
        if gt.expected_plate not in (NO_PLATE, ""):
            readable_matched += 1
            char_errors += _edit_distance(gt.expected_plate, model.plate)
            char_total += len(gt.expected_plate)

    missed = [gt for gt_i, gt in enumerate(gt_events) if gt_i not in gt_taken]
    extra = [model for model_i, model in enumerate(model_events) if model_i not in model_taken]

    tp = len(matched)
    fp = len(extra)
    fn = len(missed)
    detection = {
        "true_positive": tp,
        "false_positive": fp,
        "false_negative": fn,
        "precision": _ratio(tp, tp + fp),
        "recall": _ratio(tp, tp + fn),
    }
    recognition = {
        "matched": tp,
        "exact_matches": exact,
        "exact_accuracy_on_matched": _ratio(exact, tp),
        "end_to_end_exact_accuracy": _ratio(exact, len(gt_events)),
        "character_error_rate": _ratio(char_errors, char_total),
        "readable_matched": readable_matched,
    }
    return EvalReport(
        detection=detection,
        recognition=recognition,
        by_quality=by_quality,
        matched=matched,
        missed=missed,
        extra=extra,
    )


def report_to_dict(report: EvalReport) -> dict[str, Any]:
    return {
        "detection": report.detection,
        "recognition": report.recognition,
        "by_quality": report.by_quality,
        "missed": [
            {"index": gt.index, "expected": gt.expected_plate, "quality": gt.quality}
            for gt in report.missed
        ],
        "extra": [
            {"plate": model.plate, "start_ms": model.start_ms, "end_ms": model.end_ms}
            for model in report.extra
        ],
        "mismatches": [
            {
                "index": pair.gt.index,
                "expected": pair.gt.expected_plate,
                "model": pair.model.plate,
                "quality": pair.gt.quality,
            }
            for pair in report.matched
            if not pair.plate_correct
        ],
    }


# --------------------------------------------------------------------------- #
# Workbook parsing (needs openpyxl; separate from the metric core above).
# --------------------------------------------------------------------------- #
def _time_to_ms(text: str) -> int:
    parts = [int(p) for p in str(text).strip().split(":") if p.strip().lstrip("-").isdigit()]
    if len(parts) == 2:
        return (parts[0] * 60 + parts[1]) * 1000
    if len(parts) == 3:
        return (parts[0] * 3600 + parts[1] * 60 + parts[2]) * 1000
    if len(parts) == 1:
        return parts[0] * 1000
    return 0


def _cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def parse_qa_gt_xlsx(path: Path) -> list[GtEvent]:
    """Parse a human QA workbook (``Lane9_qa.xlsx`` layout) into GT events."""

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_index = None
    columns: dict[str, int] = {}
    for i, row in enumerate(rows):
        texts = [("" if c is None else str(c).strip().lower()) for c in row]
        if any("license plate expected" in t for t in texts):
            header_index = i
            for col_i, text in enumerate(texts):
                if "from" in text and "to" in text:
                    columns["time"] = col_i
                elif "license plate expected" in text:
                    columns["expected"] = col_i
                elif "phân loại" in text or "chất lượng" in text:
                    columns["quality"] = col_i
                elif text == "result":
                    columns["result"] = col_i
            break
    if header_index is None:
        raise ValueError("Could not locate the 'License Plate expected' header row")

    events: list[GtEvent] = []
    for row in rows[header_index + 1 :]:
        expected = normalize_plate(_cell(row, columns.get("expected")))
        time_text = _cell(row, columns.get("time"))
        if not expected or "-" not in time_text:
            continue
        start_text, _, end_text = time_text.partition("-")
        events.append(
            GtEvent(
                index=len(events) + 1,
                start_ms=_time_to_ms(start_text),
                end_ms=_time_to_ms(end_text),
                expected_plate=expected,
                quality=_cell(row, columns.get("quality")),
                human_result=_cell(row, columns.get("result")),
            )
        )
    return events


def parse_plate_report_xlsx(path: Path) -> list[ModelEvent]:
    """Parse our own ``Plate Report`` export into model events."""

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [("" if c is None else str(c).strip().lower()) for c in rows[0]]

    def find(*needles: str) -> int | None:
        for i, text in enumerate(header):
            if any(n in text for n in needles):
                return i
        return None

    plate_col = find("plate model", "plate model đọc", "model đọc")
    start_col = find("start")
    end_col = find("end")
    stt_col = find("stt")

    events: list[ModelEvent] = []
    for row in rows[1:]:
        stt = _cell(row, stt_col)
        if not stt.isdigit():
            continue
        events.append(
            ModelEvent(
                start_ms=_time_to_ms(_cell(row, start_col)),
                end_ms=_time_to_ms(_cell(row, end_col)),
                plate=normalize_plate(_cell(row, plate_col)),
            )
        )
    return events
