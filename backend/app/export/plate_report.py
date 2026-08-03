"""Native ``Plate Report`` Excel exporter.

Reproduces the human-facing report format (one row per motorcycle event) directly from
validated pipeline evidence: an embedded full frame and plate crop, the model reading,
TrackID, a data label (Phân loại), confidence, and reviewer columns.

The module is intentionally decoupled from FastAPI and SQLAlchemy: callers pass plain
``PlateReportRow`` values plus resolved evidence paths, so it can be unit-tested offline.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SHEET_TITLE = "Plate Report"
NO_PLATE_TEXT = "LPN_NO_PLATE_VEHICLE"
MISSING_IMAGE_TEXT = "Không có ảnh"

# (header, width) pairs, in reviewer order. Column indices are derived from this tuple.
_COLUMNS: tuple[tuple[str, int], ...] = (
    ("STT", 6),
    ("Ảnh frame", 70),
    ("Ảnh crop biển số", 36),
    ("TrackID", 16),
    ("Plate model đọc", 18),
    ("Phân loại", 16),
    ("GT Plate", 18),
    ("Start", 9),
    ("End", 9),
    ("Confidence", 12),
    ("Frame #", 9),
    ("Discard", 10),
    ("Kết quả QA", 18),
    ("Ghi chú QA", 32),
)
PLATE_REPORT_HEADERS: tuple[str, ...] = tuple(header for header, _ in _COLUMNS)
_COL = {header: index for index, (header, _) in enumerate(_COLUMNS, start=1)}

# Plate-quality taxonomy for the "Phân loại" data-label column. These describe a human
# judgement (glare/forged/deformed cannot be auto-classified reliably without a dedicated
# trained classifier), so the column is a dropdown the reviewer fills; only the fully
# deterministic "Xe không biển" is pre-filled from the model.
_QUALITY_OPTIONS = (
    "Biển đẹp bình thường",
    "Biển bẩn",
    "Biển cũ, xước, mờ",
    "Biển bị che vật lý",
    "Biển bị dán (icon, decal, trang trí...)",
    "Biển bị chói sáng",
    "Biển giả mạo (che 1 vài số, dán biển khác lên...)",
    "Biển biến dạng vật lý (cong, vênh)",
    "Biển bóng của vật thể che khuất",
    "Xe không biển",
)
_QUALITY_LIST_SHEET = "Lists"

_QA_OPTIONS = "Đúng,Sai,Trùng,Không biển,Ảnh mờ,Cần kiểm tra"
# Full-frame evidence is embedded large so the reviewer can actually read the scene; the crop
# a bit smaller. Both are resampled with LANCZOS + high JPEG quality so they stay sharp (the
# source frames are stored at native resolution, so the only blur was the old 200px downscale).
_FRAME_WIDTH_PX = 480
_CROP_WIDTH_PX = 240
_ROW_HEIGHT_PT = 86  # minimum row height (used when a row has no full-frame image)
# Excel row height is in points; 1 image pixel ≈ 0.75 pt. A little padding keeps a margin.
_PX_TO_PT = 0.75
_ROW_PAD_PT = 6
_CENTERED_COLUMNS = {_COL["STT"], _COL["TrackID"], _COL["Confidence"], _COL["Frame #"]}

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
_HEADER_BORDER = Border(*(Side(style="thin", color="808080"),) * 4)
_CELL_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
_CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
_CENTER_ALIGN = Alignment(vertical="center", horizontal="center")


@dataclass(frozen=True, slots=True)
class PlateReportRow:
    """A single reviewable motorcycle event."""

    plate_text: str
    start_ms: int
    end_ms: int
    frame_number: int
    confidence: float | None = None
    gt_text: str = ""
    discard: bool = False
    qa_result: str = ""
    qa_note: str = ""
    crop_path: Path | None = None
    track_code: str = ""
    classification: str = ""
    quality: str = ""  # reviewer-picked quality label; falls back to the no-plate default
    full_frame_path: Path | None = None


def format_timestamp(milliseconds: int) -> str:
    """Render milliseconds as ``m:ss`` to match the existing report."""

    total_seconds = max(0, milliseconds) // 1000
    return f"{total_seconds // 60}:{total_seconds % 60:02d}"


def format_confidence(confidence: float | None) -> str:
    return "—" if confidence is None else f"{round(confidence * 100)}%"


def quality_prefill(classification: str) -> str:
    """Only the deterministic no-plate case is pre-filled; the rest is a human dropdown."""

    return "Xe không biển" if classification == "NO_PLATE" else ""


def _scaled_image(path: Path, width_px: int, height_px: int | None = None) -> XlsxImage:
    """Resample the source to the target width (LANCZOS, high quality) and embed it."""

    from PIL import Image as PilImage

    resampling = getattr(PilImage, "Resampling", PilImage).LANCZOS
    with PilImage.open(path) as source:
        rgb = source.convert("RGB")
        if height_px is None:
            height_px = max(1, round(rgb.height * width_px / rgb.width))
        rgb = rgb.resize((width_px, height_px), resampling)
        buffer = io.BytesIO()
        rgb.save(buffer, format="JPEG", quality=90)
    buffer.seek(0)
    image = XlsxImage(buffer)
    image.width = width_px
    image.height = height_px
    return image


def build_plate_report_workbook(rows: list[PlateReportRow]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE
    sheet.freeze_panes = "A2"

    sheet.append(list(PLATE_REPORT_HEADERS))
    sheet.row_dimensions[1].height = 28
    for cell, (_, width) in zip(sheet[1], _COLUMNS, strict=True):
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER
        sheet.column_dimensions[cell.column_letter].width = width

    qa_validation = DataValidation(
        type="list", formula1=f'"{_QA_OPTIONS}"', allow_blank=True, showErrorMessage=True
    )
    sheet.add_data_validation(qa_validation)

    # The quality categories contain commas and exceed the inline-list limit, so they live
    # in a hidden sheet and the "Phân loại" dropdown references that range.
    lists = workbook.create_sheet(_QUALITY_LIST_SHEET)
    for index, option in enumerate(_QUALITY_OPTIONS, start=1):
        lists.cell(row=index, column=1, value=option)
    lists.sheet_state = "hidden"
    quality_validation = DataValidation(
        type="list",
        formula1=f"{_QUALITY_LIST_SHEET}!$A$1:$A${len(_QUALITY_OPTIONS)}",
        allow_blank=True,
        showErrorMessage=True,
    )
    sheet.add_data_validation(quality_validation)

    for position, row in enumerate(rows, start=1):
        excel_row = position + 1
        values = {
            _COL["STT"]: position,
            _COL["TrackID"]: row.track_code,
            _COL["Plate model đọc"]: row.plate_text,
            _COL["Phân loại"]: row.quality or quality_prefill(row.classification),
            _COL["GT Plate"]: row.gt_text,
            _COL["Start"]: format_timestamp(row.start_ms),
            _COL["End"]: format_timestamp(row.end_ms),
            _COL["Confidence"]: format_confidence(row.confidence),
            _COL["Frame #"]: row.frame_number,
            _COL["Discard"]: "Có" if row.discard else "Không",
            _COL["Kết quả QA"]: row.qa_result,
            _COL["Ghi chú QA"]: row.qa_note,
        }
        for column in range(1, len(_COLUMNS) + 1):
            cell = sheet.cell(row=excel_row, column=column, value=values.get(column))
            cell.border = _CELL_BORDER
            cell.alignment = _CENTER_ALIGN if column in _CENTERED_COLUMNS else _CELL_ALIGN
        qa_validation.add(sheet.cell(row=excel_row, column=_COL["Kết quả QA"]))
        quality_validation.add(sheet.cell(row=excel_row, column=_COL["Phân loại"]))

        # Track the tallest embedded image so the row grows to show the full frame in full.
        row_height_pt = _ROW_HEIGHT_PT
        if row.full_frame_path is not None and row.full_frame_path.is_file():
            frame_cell = f"{get_column_letter(_COL['Ảnh frame'])}{excel_row}"
            frame_image = _scaled_image(row.full_frame_path, _FRAME_WIDTH_PX)
            sheet.add_image(frame_image, frame_cell)
            row_height_pt = max(row_height_pt, frame_image.height * _PX_TO_PT + _ROW_PAD_PT)
        else:
            sheet.cell(row=excel_row, column=_COL["Ảnh frame"], value=MISSING_IMAGE_TEXT)

        if row.crop_path is not None and row.crop_path.is_file():
            crop_cell = f"{get_column_letter(_COL['Ảnh crop biển số'])}{excel_row}"
            crop_image = _scaled_image(row.crop_path, _CROP_WIDTH_PX)
            sheet.add_image(crop_image, crop_cell)
            row_height_pt = max(row_height_pt, crop_image.height * _PX_TO_PT + _ROW_PAD_PT)
        else:
            sheet.cell(row=excel_row, column=_COL["Ảnh crop biển số"], value=MISSING_IMAGE_TEXT)

        sheet.row_dimensions[excel_row].height = row_height_pt

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{len(rows) + 1}"
    return workbook


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
