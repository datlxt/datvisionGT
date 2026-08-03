"""GT Final (benchmark) exporter.

Produces the reusable ground-truth workbook: the human-verified truth on the left
(``License Plate expected`` + ``Phân loại`` + evidence images) and four *empty* columns
on the right for evaluating any model later. Pass/Fail/Miss/NA and Precision/Recall are
Excel formulas over the ``TP/FP/FN/NA`` column, so the metrics recompute automatically
when a reviewer scores a model run — one GT file, reusable across models.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.export.plate_report import (
    _CELL_ALIGN,
    _CELL_BORDER,
    _CENTER_ALIGN,
    _HEADER_ALIGN,
    _HEADER_BORDER,
    _HEADER_FILL,
    _HEADER_FONT,
    _PX_TO_PT,
    _QUALITY_LIST_SHEET,
    _QUALITY_OPTIONS,
    _ROW_PAD_PT,
    MISSING_IMAGE_TEXT,
    _scaled_image,
    format_timestamp,
    quality_prefill,
)

SHEET_TITLE = "GT Final"
_RESULT_OPTIONS = "TP,FP,FN,NA"

# (header, width) for the benchmark table.
_COLUMNS: tuple[tuple[str, int], ...] = (
    ("No", 6),
    ("From - To", 14),
    ("Ảnh cam toàn cảnh", 70),
    ("Ảnh biển số", 36),
    ("License Plate expected", 22),
    ("Phân loại biển số", 22),
    ("Biển số model trả về", 22),
    ("Ảnh biển số model trả về", 22),
    ("TP/FP/FN/NA", 14),
    ("Note", 30),
)
GT_FINAL_HEADERS: tuple[str, ...] = tuple(header for header, _ in _COLUMNS)
_COL = {header: index for index, (header, _) in enumerate(_COLUMNS, start=1)}

# Full-frame evidence embedded large (LANCZOS + high JPEG quality via _scaled_image) so the
# reviewer can read the scene; the plate crop a bit smaller. Row height grows to fit them.
_FRAME_WIDTH_PX = 480
_CROP_WIDTH_PX = 240
_ROW_HEIGHT_PT = 64  # minimum row height (used when a row has no full-frame image)
_HEADER_ROW = 11  # summary block occupies rows 1..9


@dataclass(frozen=True, slots=True)
class GtFinalRow:
    gt_text: str
    classification: str
    start_ms: int
    end_ms: int
    quality: str = ""  # reviewer-picked quality label; falls back to the no-plate default
    full_frame_path: Path | None = None
    crop_path: Path | None = None


def _summary_block(sheet, result_range: str) -> None:
    rows = [
        ("BẢNG KẾT QUẢ", None),
        ("Chỉ số", "Số lượng"),
        ("Pass (TP - nhận diện đúng)", f'=COUNTIF({result_range},"TP")'),
        ("Fail (FP - nhận diện sai)", f'=COUNTIF({result_range},"FP")'),
        ("Miss (FN - bỏ sót)", f'=COUNTIF({result_range},"FN")'),
        ("NA (Lặp đúng)", f'=COUNTIF({result_range},"NA")'),
        ("Tổng đánh giá (Pass+Fail+Miss)", "=C3+C4+C5"),
        ("Precision = TP/(TP+FP)", '=IF((C3+C4)=0,"",C3/(C3+C4))'),
        ("Recall = TP/(TP+FN)", '=IF((C3+C5)=0,"",C3/(C3+C5))'),
    ]
    for index, (label, value) in enumerate(rows, start=1):
        label_cell = sheet.cell(row=index, column=2, value=label)
        if index in (1, 2):
            label_cell.font = _HEADER_FONT
        if value is not None:
            value_cell = sheet.cell(row=index, column=3, value=value)
            if index in (8, 9):  # Precision / Recall as percentages
                value_cell.number_format = "0.0%"


def build_gt_final_workbook(rows: list[GtFinalRow]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    last_data_row = _HEADER_ROW + max(len(rows), 1)
    result_col = get_column_letter(_COL["TP/FP/FN/NA"])
    _summary_block(sheet, f"${result_col}${_HEADER_ROW + 1}:${result_col}${last_data_row}")

    for index, (header, width) in enumerate(_COLUMNS, start=1):
        cell = sheet.cell(row=_HEADER_ROW, column=index, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _HEADER_BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[_HEADER_ROW].height = 28
    sheet.freeze_panes = f"A{_HEADER_ROW + 1}"

    # Quality dropdown (reuse the plate-report taxonomy on a hidden sheet).
    lists = workbook.create_sheet(_QUALITY_LIST_SHEET)
    for index, option in enumerate(_QUALITY_OPTIONS, start=1):
        lists.cell(row=index, column=1, value=option)
    lists.sheet_state = "hidden"
    quality_validation = DataValidation(
        type="list",
        formula1=f"{_QUALITY_LIST_SHEET}!$A$1:$A${len(_QUALITY_OPTIONS)}",
        allow_blank=True,
    )
    result_validation = DataValidation(
        type="list", formula1=f'"{_RESULT_OPTIONS}"', allow_blank=True
    )
    sheet.add_data_validation(quality_validation)
    sheet.add_data_validation(result_validation)

    for position, row in enumerate(rows, start=1):
        excel_row = _HEADER_ROW + position
        values = {
            _COL["No"]: position,
            _COL["From - To"]: f"{format_timestamp(row.start_ms)} - {format_timestamp(row.end_ms)}",
            _COL["License Plate expected"]: row.gt_text,
            _COL["Phân loại biển số"]: row.quality or quality_prefill(row.classification),
        }
        for column in range(1, len(_COLUMNS) + 1):
            cell = sheet.cell(row=excel_row, column=column, value=values.get(column))
            cell.border = _CELL_BORDER
            cell.alignment = _CENTER_ALIGN if column in (1, 2, 9) else _CELL_ALIGN
        quality_validation.add(sheet.cell(row=excel_row, column=_COL["Phân loại biển số"]))
        result_validation.add(sheet.cell(row=excel_row, column=_COL["TP/FP/FN/NA"]))

        # Grow the row to fit the tallest embedded image so the full frame shows in full.
        row_height_pt = _ROW_HEIGHT_PT
        if row.full_frame_path is not None and row.full_frame_path.is_file():
            anchor = f"{get_column_letter(_COL['Ảnh cam toàn cảnh'])}{excel_row}"
            frame_image = _scaled_image(row.full_frame_path, _FRAME_WIDTH_PX)
            sheet.add_image(frame_image, anchor)
            row_height_pt = max(row_height_pt, frame_image.height * _PX_TO_PT + _ROW_PAD_PT)
        else:
            sheet.cell(row=excel_row, column=_COL["Ảnh cam toàn cảnh"], value=MISSING_IMAGE_TEXT)

        if row.crop_path is not None and row.crop_path.is_file():
            anchor = f"{get_column_letter(_COL['Ảnh biển số'])}{excel_row}"
            crop_image = _scaled_image(row.crop_path, _CROP_WIDTH_PX)
            sheet.add_image(crop_image, anchor)
            row_height_pt = max(row_height_pt, crop_image.height * _PX_TO_PT + _ROW_PAD_PT)
        else:
            sheet.cell(row=excel_row, column=_COL["Ảnh biển số"], value=MISSING_IMAGE_TEXT)

        sheet.row_dimensions[excel_row].height = row_height_pt

    return workbook
