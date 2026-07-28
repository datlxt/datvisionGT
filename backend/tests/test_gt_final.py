from __future__ import annotations

import io

from openpyxl import load_workbook

from app.export import GT_FINAL_HEADERS, GtFinalRow, build_gt_final_workbook, workbook_to_bytes


def test_gt_final_benchmark_layout_and_formulas() -> None:
    rows = [
        GtFinalRow(gt_text="29N196452", classification="RECOGNIZED", start_ms=3_000, end_ms=13_000),
        GtFinalRow(
            gt_text="", classification="NO_PLATE", start_ms=20_000, end_ms=21_000
        ),
    ]
    workbook = load_workbook(io.BytesIO(workbook_to_bytes(build_gt_final_workbook(rows))))
    sheet = workbook.active

    assert sheet.title == "GT Final"
    # Summary block with auto-metric formulas over the TP/FP/FN/NA column (I).
    assert sheet["B1"].value == "BẢNG KẾT QUẢ"
    assert sheet["B3"].value.startswith("Pass")
    assert sheet["C3"].value == '=COUNTIF($I$12:$I$13,"TP")'
    assert sheet["C8"].value == '=IF((C3+C4)=0,"",C3/(C3+C4))'  # Precision
    assert sheet["C9"].value == '=IF((C3+C5)=0,"",C3/(C3+C5))'  # Recall

    # Benchmark header at row 11.
    assert tuple(cell.value for cell in sheet[11]) == GT_FINAL_HEADERS

    # GT (left) filled, model columns (right) blank.
    assert sheet["B12"].value == "0:03 - 0:13"
    assert sheet["E12"].value == "29N196452"  # License Plate expected
    assert sheet["F12"].value in (None, "")  # Phân loại: dropdown
    assert sheet["G12"].value is None  # Ảnh biển số model trả về (để trống)
    assert sheet["H12"].value is None  # Biển số model trả về
    assert sheet["I12"].value is None  # TP/FP/FN/NA
    assert sheet["J12"].value is None  # Note
    # No-plate quality pre-fill.
    assert sheet["F13"].value == "Xe không biển"


def test_gt_final_empty_still_valid() -> None:
    workbook = load_workbook(io.BytesIO(workbook_to_bytes(build_gt_final_workbook([]))))
    sheet = workbook.active
    assert tuple(cell.value for cell in sheet[11]) == GT_FINAL_HEADERS
