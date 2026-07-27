from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from app.export import (
    PLATE_REPORT_HEADERS,
    PlateReportRow,
    build_plate_report_workbook,
    workbook_to_bytes,
)
from app.export.plate_report import (
    format_confidence,
    format_timestamp,
    quality_prefill,
)


def test_timestamp_and_confidence_formatting_match_report() -> None:
    assert format_timestamp(3_000) == "0:03"
    assert format_timestamp(124_000) == "2:04"
    assert format_timestamp(-5) == "0:00"
    assert format_confidence(0.97) == "97%"
    assert format_confidence(None) == "—"


def test_quality_prefill_only_fills_no_plate() -> None:
    # Quality is a human judgement; only the deterministic no-plate case is pre-filled.
    assert quality_prefill("NO_PLATE") == "Xe không biển"
    assert quality_prefill("RECOGNIZED") == ""
    assert quality_prefill("UNREADABLE") == ""
    assert quality_prefill("") == ""


def _make_image(path: Path, size: tuple[int, int]) -> Path:
    Image.new("RGB", size, color=(40, 80, 160)).save(path, format="JPEG")
    return path


def test_workbook_layout_values_and_embedded_media(tmp_path: Path) -> None:
    crop = _make_image(tmp_path / "plate.jpg", (120, 60))
    frame = _make_image(tmp_path / "frame.jpg", (1280, 720))
    rows = [
        PlateReportRow(
            plate_text="29N196452",
            start_ms=3_000,
            end_ms=13_000,
            frame_number=44,
            confidence=0.97,
            crop_path=crop,
            full_frame_path=frame,
            track_code="VEHICLE_000002",
            classification="RECOGNIZED",
        ),
        PlateReportRow(
            plate_text="LPN_NO_PLATE_VEHICLE",
            start_ms=139_000,
            end_ms=140_000,
            frame_number=6,
            confidence=None,
            crop_path=None,
            full_frame_path=frame,
            track_code="VEHICLE_000040",
            classification="NO_PLATE",
        ),
    ]

    workbook = load_workbook(io.BytesIO(workbook_to_bytes(build_plate_report_workbook(rows))))
    sheet = workbook.active

    assert sheet.title == "Plate Report"
    assert tuple(cell.value for cell in sheet[1]) == PLATE_REPORT_HEADERS

    # Recognized row: TrackID (D), plate (E), Phân loại (F), timings, confidence.
    assert sheet["A2"].value == 1
    assert sheet["D2"].value == "VEHICLE_000002"
    assert sheet["E2"].value == "29N196452"
    assert sheet["F2"].value in (None, "")  # quality is a dropdown the reviewer fills
    assert sheet["G2"].value is None  # GT Plate stays empty for the reviewer.
    assert sheet["H2"].value == "0:03"
    assert sheet["I2"].value == "0:13"
    assert sheet["J2"].value == "97%"
    assert sheet["K2"].value == 44

    # NO_PLATE row: sentinel text, no crop image, "Xe không biển" label.
    assert sheet["E3"].value == "LPN_NO_PLATE_VEHICLE"
    assert sheet["F3"].value == "Xe không biển"
    assert sheet["J3"].value == "—"
    assert sheet["C3"].value == "Không có ảnh"

    # Full frame embedded on both rows; crop only on the recognized row (2 + 1 = 3).
    assert len(sheet._images) == 3
    assert sheet.auto_filter.ref == "A1:N3"


def test_empty_report_has_only_headers() -> None:
    workbook = load_workbook(io.BytesIO(workbook_to_bytes(build_plate_report_workbook([]))))
    sheet = workbook.active
    assert sheet.max_row == 1
    assert tuple(cell.value for cell in sheet[1]) == PLATE_REPORT_HEADERS
